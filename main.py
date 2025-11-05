#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
期刊术语提取工具 - 主程序
功能：
1. 智能PDF文本提取（文本型直接提取，扫描版使用科大讯飞OCR）
2. 识别文献中的"关键词"和"key words"两栏
3. 将识别到的中英关键词对应起来，形成术语对，输出为excel表格
4. 同时识别"摘要"和"abstract"，发送给GPT-4o抽取术语对，输出到另一份excel中

支持两种运行模式：
- 交互式界面：直接运行此脚本（双击或 python main.py）
- 命令行参数：python main.py <PDF文件> [--output 输出目录] [--api-key API密钥]
"""

import os
import sys
import re
import json
import logging
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any

# 禁用字节码缓存，确保代码修改立即生效
sys.dont_write_bytecode = True

# 设置控制台编码为UTF-8（Windows兼容性）
if sys.platform == 'win32':
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleOutputCP(65001)
        kernel32.SetConsoleCP(65001)
    except:
        pass

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('journal_extractor.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =============================================================================
# 依赖检查和导入
# =============================================================================

missing_modules = []

try:
    from file_processor import FileProcessor
except ImportError:
    missing_modules.append("file_processor.py")
    FileProcessor = None

try:
    from gpt_processor import GPTProcessor
except ImportError:
    missing_modules.append("gpt_processor.py")
    GPTProcessor = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    missing_modules.append("openpyxl (pip install openpyxl)")
    openpyxl = None

if missing_modules:
    print(f"❌ 缺少以下模块:")
    for module in missing_modules:
        print(f"   - {module}")
    print("\n请安装所需依赖:")
    print("pip install requests openpyxl openai tiktoken pdfminer.six PyPDF2")
    sys.exit(1)


# =============================================================================
# 核心功能类：期刊关键词和摘要提取器
# =============================================================================

class JournalKeywordExtractor:
    """期刊关键词和摘要提取器"""
    
    def __init__(self, api_key: str = None, base_url: str = None):
        """
        初始化提取器
        
        Args:
            api_key: OpenAI API密钥（如果不提供，会从config.py或环境变量读取）
            base_url: API端点URL（如果不提供，会从config.py或环境变量读取）
        """
        # 尝试从config.py读取配置
        try:
            from config import OPENAI_API_KEY, OPENAI_BASE_URL
            self.api_key = api_key or os.getenv("OPENAI_API_KEY") or OPENAI_API_KEY
            self.base_url = base_url or os.getenv("OPENAI_BASE_URL") or OPENAI_BASE_URL
        except ImportError:
            self.api_key = api_key or os.getenv("OPENAI_API_KEY")
            self.base_url = base_url or os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
        
        # 初始化文件处理器（智能PDF提取）
        self.file_processor = FileProcessor(use_gpu=False, enable_ocr=True)
        
        # 初始化GPT处理器（如果提供了API密钥）
        self.gpt_processor = None
        if self.api_key:
            self.gpt_processor = GPTProcessor(
                api_key=self.api_key,
                base_url=self.base_url,
                enable_checkpoint=False
            )
        
        logger.info("JournalKeywordExtractor初始化完成")
    
    # =============================================================================
    # PDF文本提取
    # =============================================================================
    
    def extract_text_from_pdf(self, pdf_path: str, save_text: bool = True, output_dir: Path = None) -> str:
        """
        智能提取PDF文本（文本型直接提取，扫描版使用科大讯飞OCR）
        
        Args:
            pdf_path: PDF文件路径
            save_text: 是否保存提取的文本到文件
            output_dir: 输出目录（用于保存文本）
            
        Returns:
            str: 提取的完整文本
        """
        logger.info(f"开始提取PDF文本: {pdf_path}")
        print(f"\n" + "=" * 70)
        print(f"📄 步骤1/5: 提取PDF文本")
        print("=" * 70)
        print(f"文件: {Path(pdf_path).name}")
        
        try:
            # 使用文件处理器提取文本（自动判断文本型/扫描版）
            file_type, texts = self.file_processor.process_file(pdf_path)
            
            # 保存提取的文本到文件
            if save_text and output_dir:
                pdf_name = Path(pdf_path).stem
                ocr_dir = output_dir / "ocr_extracted_text"
                ocr_dir.mkdir(exist_ok=True)
                text_file = ocr_dir / f"{pdf_name}_OCR提取文本.txt"
                
                print(f"💾 保存提取的文本到: {text_file.name}")
                self.file_processor.save_extracted_text(texts, str(text_file))
                print(f"✅ 文本已保存")
            
            # 合并所有文本
            full_text = '\n\n'.join(texts)
            
            logger.info(f"成功提取PDF文本，共{len(full_text)}字符")
            print(f"✅ PDF文本提取成功，共{len(full_text):,}字符")
            print("=" * 70)
            
            return full_text
            
        except Exception as e:
            logger.error(f"PDF文本提取失败: {e}")
            print(f"❌ PDF文本提取失败: {e}")
            raise
    
    # =============================================================================
    # 关键词和摘要识别
    # =============================================================================
    
    def extract_all_keywords_from_journal(self, text: str) -> List[Tuple[Optional[str], Optional[str]]]:
        """
        从整个期刊中提取所有文章的关键词
        
        Args:
            text: 完整文本
            
        Returns:
            List[Tuple[Optional[str], Optional[str]]]: 所有文章的(中文关键词, 英文关键词)列表
        """
        logger.info("开始识别所有文章的关键词...")
        print(f"🔍 正在搜索所有文章的关键词...")
        
        all_keywords = []
        
        # 改进的正则表达式 - 更精确地匹配关键词
        # 中文关键词：匹配到中图分类号、DOI、文献标识码等标记
        zh_pattern = r'关键词\s*[：:]\s*(.*?)(?=\s*(?:中图分类号|DOI|文献标识码|Key\s*words?|引用格式|\n\n\n|$))'
        
        # 英文关键词：匹配到下一个段落分隔或常见标记
        en_pattern = r'Key\s*words?\s*[：:]\s*(.*?)(?=\s*(?:\n\n|1\s+引|收稿日期|基金项目|作者简介|$))'
        
        # 查找所有中文关键词
        zh_matches = list(re.finditer(zh_pattern, text, re.IGNORECASE | re.DOTALL))
        logger.info(f"找到{len(zh_matches)}处中文关键词")
        
        # 查找所有英文关键词
        en_matches = list(re.finditer(en_pattern, text, re.IGNORECASE | re.DOTALL))
        logger.info(f"找到{len(en_matches)}处英文关键词")
        
        # 配对中英文关键词
        max_matches = max(len(zh_matches), len(en_matches))
        
        for i in range(max_matches):
            zh_text = None
            en_text = None
            
            if i < len(zh_matches):
                zh_text = zh_matches[i].group(1).strip()
                # 清理不需要的内容
                zh_text = re.sub(r'\s+', ' ', zh_text)  # 规范化空格
                
            if i < len(en_matches):
                en_text = en_matches[i].group(1).strip()
                # 清理不需要的内容
                en_text = re.sub(r'\s+', ' ', en_text)
            
            if zh_text or en_text:
                all_keywords.append((zh_text, en_text))
        
        print(f"✅ 找到{len(all_keywords)}篇文章的关键词")
        return all_keywords
    
    def extract_keywords_section(self, text: str) -> Tuple[Optional[str], Optional[str]]:
        """
        提取关键词部分（中文和英文）- 兼容旧接口
        
        Args:
            text: 完整文本
            
        Returns:
            Tuple[Optional[str], Optional[str]]: (中文关键词文本, 英文关键词文本)
        """
        all_keywords = self.extract_all_keywords_from_journal(text)
        
        if all_keywords:
            # 如果有多篇文章，合并所有关键词
            zh_all = []
            en_all = []
            
            for zh, en in all_keywords:
                if zh:
                    zh_all.append(zh)
                if en:
                    en_all.append(en)
            
            zh_combined = ' | '.join(zh_all) if zh_all else None
            en_combined = ' | '.join(en_all) if en_all else None
            
            return zh_combined, en_combined
        else:
            logger.warning("未找到明确的关键词部分")
            print("⚠️  未找到明确的关键词部分")
            return None, None
    
    def extract_all_abstracts_from_journal(self, text: str) -> List[Tuple[Optional[str], Optional[str]]]:
        """
        从整个期刊中提取所有文章的摘要
        
        Args:
            text: 完整文本
            
        Returns:
            List[Tuple[Optional[str], Optional[str]]]: 所有文章的(中文摘要, 英文摘要)列表
        """
        logger.info("开始识别所有文章的摘要...")
        print(f"🔍 正在搜索所有文章的摘要...")
        
        all_abstracts = []
        
        # 改进的正则表达式 - 更精确地匹配摘要
        # 中文摘要：匹配到关键词标记
        zh_pattern = r'摘\s*要\s*[：:]\s*(.*?)(?=\s*(?:关键词|关键字))'
        
        # 英文摘要：匹配到Key words标记
        en_pattern = r'Abstract\s*[：:]\s*(.*?)(?=\s*(?:Key\s*words?|KEYWORDS?))'
        
        # 查找所有中文摘要
        zh_matches = list(re.finditer(zh_pattern, text, re.IGNORECASE | re.DOTALL))
        logger.info(f"找到{len(zh_matches)}处中文摘要")
        
        # 查找所有英文摘要
        en_matches = list(re.finditer(en_pattern, text, re.IGNORECASE | re.DOTALL))
        logger.info(f"找到{len(en_matches)}处英文摘要")
        
        # 配对中英文摘要
        max_matches = max(len(zh_matches), len(en_matches))
        
        for i in range(max_matches):
            zh_text = None
            en_text = None
            
            if i < len(zh_matches):
                zh_text = zh_matches[i].group(1).strip()
                # 规范化空格
                zh_text = re.sub(r'\s+', ' ', zh_text)
                
            if i < len(en_matches):
                en_text = en_matches[i].group(1).strip()
                # 规范化空格
                en_text = re.sub(r'\s+', ' ', en_text)
            
            if zh_text or en_text:
                all_abstracts.append((zh_text, en_text))
        
        print(f"✅ 找到{len(all_abstracts)}篇文章的摘要")
        return all_abstracts
    
    def extract_abstract_section(self, text: str) -> Tuple[Optional[str], Optional[str]]:
        """
        提取摘要部分（中文和英文）- 兼容旧接口
        
        Args:
            text: 完整文本
            
        Returns:
            Tuple[Optional[str], Optional[str]]: (中文摘要文本, 英文摘要文本)
        """
        all_abstracts = self.extract_all_abstracts_from_journal(text)
        
        if all_abstracts:
            # 合并所有摘要（用于GPT提取）
            zh_all = []
            en_all = []
            
            for zh, en in all_abstracts:
                if zh:
                    zh_all.append(zh)
                if en:
                    en_all.append(en)
            
            zh_combined = '\n\n---\n\n'.join(zh_all) if zh_all else None
            en_combined = '\n\n---\n\n'.join(en_all) if en_all else None
            
            logger.info(f"找到摘要部分 - 中文: {len(zh_all)}篇, 英文: {len(en_all)}篇")
            print(f"✅ 找到{len(all_abstracts)}篇文章的摘要")
            if zh_combined:
                print(f"   中文摘要总计: {len(zh_combined)}字符")
            if en_combined:
                print(f"   英文摘要总计: {len(en_combined)}字符")
            
            return zh_combined, en_combined
        else:
            logger.warning("未找到明确的摘要部分")
            print("⚠️  未找到明确的摘要部分")
            return None, None
    
    # =============================================================================
    # 关键词解析和匹配
    # =============================================================================
    
    def parse_keywords(self, keywords_text: str) -> List[str]:
        """
        解析关键词字符串，分割成单个关键词
        
        Args:
            keywords_text: 关键词文本
            
        Returns:
            List[str]: 关键词列表
        """
        if not keywords_text:
            return []
        
        # 清理文本（移除首尾空格和换行）
        keywords_text = keywords_text.strip()
        keywords_text = re.sub(r'\s+', ' ', keywords_text)  # 规范化空格
        
        # 尝试多种分隔符（按优先级）
        separators = [';', '；', ',', '，', '、']
        
        keywords = []
        for sep in separators:
            if sep in keywords_text:
                parts = keywords_text.split(sep)
                keywords = [k.strip() for k in parts if k.strip()]
                if keywords:  # 找到有效分割就停止
                    break
        
        # 如果没有找到分隔符，尝试按多个空格分割（针对英文）
        if not keywords:
            # 检查是否有多个连续空格（可能是列表格式）
            if '  ' in keywords_text or '\t' in keywords_text:
                keywords = [k.strip() for k in re.split(r'\s{2,}|\t', keywords_text) if k.strip()]
        
        # 如果还是没有，整个文本作为一个关键词
        if not keywords:
            keywords = [keywords_text]
        
        # 清理每个关键词中的多余空格和特殊字符
        keywords = [re.sub(r'\s+', ' ', k.strip()) for k in keywords]
        
        return keywords
    
    def match_bilingual_keywords(self, zh_keywords: List[str], en_keywords: List[str]) -> List[Dict[str, str]]:
        """
        匹配中英文关键词对
        
        Args:
            zh_keywords: 中文关键词列表
            en_keywords: 英文关键词列表
            
        Returns:
            List[Dict[str, str]]: 关键词对列表，每个字典包含'zh_term'和'en_term'
        """
        logger.info(f"开始匹配关键词对: 中文{len(zh_keywords)}个, 英文{len(en_keywords)}个")
        
        keyword_pairs = []
        
        # 方法1: 按顺序匹配（假设顺序对应）
        max_len = max(len(zh_keywords), len(en_keywords))
        
        for i in range(max_len):
            zh_term = zh_keywords[i] if i < len(zh_keywords) else ""
            en_term = en_keywords[i] if i < len(en_keywords) else ""
            
            if zh_term or en_term:
                keyword_pairs.append({
                    'zh_term': zh_term,
                    'en_term': en_term
                })
        
        logger.info(f"成功匹配{len(keyword_pairs)}个关键词对")
        return keyword_pairs
    
    # =============================================================================
    # GPT术语提取
    # =============================================================================
    
    def extract_terms_from_abstracts_with_gpt(self, abstracts_list: List[Tuple[Optional[str], Optional[str]]]) -> List[Dict[str, str]]:
        """
        使用GPT-4o从多篇文章的摘要中分别提取术语对
        
        Args:
            abstracts_list: 摘要列表，每个元素为(中文摘要, 英文摘要)
            
        Returns:
            List[Dict[str, str]]: 所有术语对列表
        """
        if not self.gpt_processor:
            raise ValueError("未配置GPT处理器，请提供OpenAI API密钥")
        
        logger.info(f"开始使用GPT-4o提取{len(abstracts_list)}篇文章的摘要术语...")
        print(f"\n🤖 正在使用GPT-4o分析摘要并提取术语对...")
        print(f"   将分别处理{len(abstracts_list)}篇文章的摘要")
        print("=" * 70)
        
        all_term_pairs = []
        
        # 尝试从config.py读取提示词配置
        try:
            from config import GPT_SYSTEM_PROMPT, GPT_USER_PROMPT_TEMPLATE
            system_prompt = GPT_SYSTEM_PROMPT
            user_prompt_template = GPT_USER_PROMPT_TEMPLATE
        except ImportError:
            system_prompt = """You are a technical terminology extraction specialist. Your task is to extract bilingual technical term pairs from academic abstracts."""
            user_prompt_template = """Please extract technical term pairs from the following abstract.

For each technical term you identify, provide both its English and Chinese versions.

EXTRACTION RULES:
1. Focus on domain-specific technical terms, concepts, and specialized vocabulary
2. EVERY term must have BOTH English and Chinese versions
3. If abstract is in Chinese, provide English translation
4. If abstract is in English, provide Chinese translation
5. Extract 5-10 most important technical terms from THIS abstract
6. Avoid generic words unless they have specific technical meaning

OUTPUT FORMAT (strict JSON):
{{
  "terms": [
    {{
      "en_term": "English technical term",
      "zh_term": "对应的中文术语"
    }}
  ]
}}

ABSTRACT TEXT:
{abstract_text}

Please extract the technical terms now:"""
        
        # 逐篇处理摘要
        for article_idx, (zh_abstract, en_abstract) in enumerate(abstracts_list, 1):
            try:
                print(f"\n   处理第{article_idx}/{len(abstracts_list)}篇文章...")
                
                # 构建提示词
                combined_abstract = ""
                if zh_abstract:
                    combined_abstract += f"【中文摘要】\n{zh_abstract}\n\n"
                if en_abstract:
                    combined_abstract += f"【English Abstract】\n{en_abstract}"
                
                if not combined_abstract.strip():
                    print(f"   ⚠️  第{article_idx}篇摘要为空，跳过")
                    continue
                
                user_prompt = user_prompt_template.format(abstract_text=combined_abstract)
                
                # 调用GPT处理
                result = self.gpt_processor.process_single_text(
                    text=combined_abstract,
                    custom_id=f"abstract_term_extraction_article_{article_idx}",
                    system_prompt=system_prompt,
                    user_prompt_template=user_prompt,
                    model="gpt-4o",
                    temperature=0,
                    max_tokens=2048
                )
                
                # 解析结果
                extracted_terms = result.get('extracted_terms', {})
                
                if 'terms' in extracted_terms and isinstance(extracted_terms['terms'], list):
                    term_pairs = extracted_terms['terms']
                    # 验证术语对格式
                    valid_pairs = []
                    for pair in term_pairs:
                        if isinstance(pair, dict) and 'en_term' in pair and 'zh_term' in pair:
                            valid_pairs.append(pair)
                        else:
                            logger.warning(f"跳过格式不正确的术语对: {pair}")
                    
                    if valid_pairs:
                        all_term_pairs.extend(valid_pairs)
                        print(f"   ✅ 第{article_idx}篇提取了{len(valid_pairs)}个术语对")
                    else:
                        print(f"   ⚠️  第{article_idx}篇术语对格式不正确")
                else:
                    print(f"   ⚠️  第{article_idx}篇GPT返回格式不正确")
                
                # 添加延迟避免API限流
                if article_idx < len(abstracts_list):
                    import time
                    time.sleep(1)
                    
            except KeyboardInterrupt:
                print("\n\n⏹️  用户中断GPT处理")
                raise
            except Exception as e:
                logger.error(f"第{article_idx}篇GPT术语提取失败: {e}")
                print(f"   ❌ 第{article_idx}篇处理失败: {e}")
                continue
        
        logger.info(f"GPT总共提取了{len(all_term_pairs)}个术语对")
        print(f"\n✅ GPT总共提取了{len(all_term_pairs)}个术语对（来自{len(abstracts_list)}篇文章）")
        print("=" * 70)
        
        return all_term_pairs
    
    # =============================================================================
    # Excel输出
    # =============================================================================
    
    def save_keywords_to_excel(self, keyword_pairs: List[Dict[str, str]], output_file: str):
        """
        将关键词对保存为Excel文件
        
        Args:
            keyword_pairs: 关键词对列表
            output_file: 输出文件路径
        """
        logger.info(f"开始保存关键词到Excel: {output_file}")
        print(f"\n💾 正在保存关键词到Excel...")
        
        try:
            if not openpyxl:
                raise ImportError("openpyxl未安装，请运行: pip install openpyxl")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "关键词术语对"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # 写入标题行
            headers = ["序号", "中文关键词", "英文关键词", "来源"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 写入数据
            for idx, pair in enumerate(keyword_pairs, 1):
                row_data = [
                    idx,
                    pair.get('zh_term', ''),
                    pair.get('en_term', ''),
                    '文献关键词'
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=idx + 1, column=col, value=value)
                    cell.border = border
                    if col == 1:
                        cell.alignment = Alignment(horizontal="center")
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 15
            
            # 保存文件
            wb.save(output_file)
            logger.info(f"关键词Excel保存成功: {output_file}")
            print(f"✅ 关键词Excel已保存: {output_file}")
            
        except Exception as e:
            logger.error(f"保存Excel失败: {e}")
            print(f"❌ 保存Excel失败: {e}")
            raise
    
    def save_abstract_terms_to_excel(self, term_pairs: List[Dict[str, str]], output_file: str):
        """
        将从摘要提取的术语对保存为Excel文件
        
        Args:
            term_pairs: 术语对列表
            output_file: 输出文件路径
        """
        logger.info(f"开始保存摘要术语到Excel: {output_file}")
        print(f"\n💾 正在保存摘要术语到Excel...")
        
        try:
            if not openpyxl:
                raise ImportError("openpyxl未安装，请运行: pip install openpyxl")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "摘要术语对"
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # 写入标题行
            headers = ["序号", "中文术语", "英文术语", "来源"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 写入数据
            for idx, pair in enumerate(term_pairs, 1):
                row_data = [
                    idx,
                    pair.get('zh_term', ''),
                    pair.get('en_term', ''),
                    'GPT-4o摘要提取'
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=idx + 1, column=col, value=value)
                    cell.border = border
                    if col == 1:
                        cell.alignment = Alignment(horizontal="center")
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 20
            
            # 保存文件
            wb.save(output_file)
            logger.info(f"摘要术语Excel保存成功: {output_file}")
            print(f"✅ 摘要术语Excel已保存: {output_file}")
            
        except Exception as e:
            logger.error(f"保存Excel失败: {e}")
            print(f"❌ 保存Excel失败: {e}")
            raise
    
    # =============================================================================
    # 完整处理流程
    # =============================================================================
    
    def process_journal_pdf(self, pdf_path: str, output_dir: str = None) -> Dict[str, str]:
        """
        处理期刊PDF文件，提取关键词和摘要术语
        
        Args:
            pdf_path: PDF文件路径
            output_dir: 输出目录（默认为当前目录）
            
        Returns:
            Dict[str, str]: 包含输出文件路径的字典
        """
        print("\n" + "=" * 70)
        print("🎯 期刊术语提取工具")
        print("=" * 70)
        
        # 确定输出目录
        if output_dir is None:
            output_dir = Path(pdf_path).parent
        else:
            output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
        
        # 生成输出文件名
        pdf_name = Path(pdf_path).stem
        keywords_output = output_dir / f"{pdf_name}_关键词术语对.xlsx"
        abstract_output = output_dir / f"{pdf_name}_摘要术语对.xlsx"
        
        results = {}
        
        try:
            # 检查文件是否存在
            if not os.path.exists(pdf_path):
                raise FileNotFoundError(f"PDF文件不存在: {pdf_path}")
            
            # 步骤1: 提取PDF文本
            full_text = self.extract_text_from_pdf(pdf_path, save_text=True, output_dir=output_dir)
            
            if not full_text or len(full_text.strip()) < 100:
                raise ValueError("PDF文本提取失败或内容过少")
            
            # 步骤2: 提取所有文章的关键词
            print("\n" + "=" * 70)
            print("📑 步骤2/5: 识别关键词部分")
            print("=" * 70)
            all_keywords_pairs = self.extract_all_keywords_from_journal(full_text)
            
            # 步骤3: 解析和匹配关键词
            if all_keywords_pairs:
                print("\n" + "=" * 70)
                print("🔗 步骤3/5: 解析和匹配关键词")
                print("=" * 70)
                
                # 对每篇文章的关键词进行解析和匹配
                all_term_pairs = []
                
                for article_idx, (zh_text, en_text) in enumerate(all_keywords_pairs, 1):
                    print(f"\n   处理第{article_idx}篇文章的关键词...")
                    
                    # 解析中英文关键词
                    zh_keywords = self.parse_keywords(zh_text) if zh_text else []
                    en_keywords = self.parse_keywords(en_text) if en_text else []
                    
                    print(f"   - 中文关键词: {len(zh_keywords)}个")
                    print(f"   - 英文关键词: {len(en_keywords)}个")
                    
                    # 匹配关键词对
                    article_pairs = self.match_bilingual_keywords(zh_keywords, en_keywords)
                    all_term_pairs.extend(article_pairs)
                
                print(f"\n✅ 总计提取了{len(all_term_pairs)}个关键词术语对")
                
                # 保存所有关键词到Excel
                self.save_keywords_to_excel(all_term_pairs, str(keywords_output))
                results['keywords_file'] = str(keywords_output)
            else:
                print("\n⚠️  未找到关键词，跳过关键词提取")
            
            # 步骤4: 提取所有文章的摘要
            print("\n" + "=" * 70)
            print("📝 步骤4/5: 识别摘要部分")
            print("=" * 70)
            all_abstracts = self.extract_all_abstracts_from_journal(full_text)
            
            # 合并所有摘要用于GPT处理
            if all_abstracts:
                zh_abstracts = [zh for zh, en in all_abstracts if zh]
                en_abstracts = [en for zh, en in all_abstracts if en]
                zh_abstract = '\n\n---文章分隔---\n\n'.join(zh_abstracts) if zh_abstracts else None
                en_abstract = '\n\n---Article Separator---\n\n'.join(en_abstracts) if en_abstracts else None
            else:
                zh_abstract = None
                en_abstract = None
            
            # 步骤5: 使用GPT提取摘要术语
            if all_abstracts and self.gpt_processor:
                print("\n" + "=" * 70)
                print("🤖 步骤5/5: GPT-4o分析摘要提取术语")
                print("=" * 70)
                
                abstract_terms = self.extract_terms_from_abstracts_with_gpt(all_abstracts)
                
                if abstract_terms:
                    # 保存摘要术语到Excel
                    self.save_abstract_terms_to_excel(abstract_terms, str(abstract_output))
                    results['abstract_terms_file'] = str(abstract_output)
                else:
                    print("⚠️  GPT未提取到术语")
            elif not all_abstracts:
                print("\n⚠️  未找到摘要，跳过GPT术语提取")
            elif not self.gpt_processor:
                print("\n⚠️  未配置GPT API，跳过摘要术语提取")
            
            # 最终总结
            print("\n" + "=" * 70)
            print("🎉 处理完成！")
            print("=" * 70)
            if results.get('keywords_file'):
                print(f"📄 关键词术语对: {Path(results['keywords_file']).name}")
            if results.get('abstract_terms_file'):
                print(f"📄 摘要术语对: {Path(results['abstract_terms_file']).name}")
            print("=" * 70 + "\n")
            
            logger.info("处理完成")
            return results
            
        except Exception as e:
            logger.error(f"处理PDF失败: {e}")
            print(f"\n❌ 处理失败: {e}")
            raise


# =============================================================================
# 交互式界面类
# =============================================================================

class InteractiveTermExtractor:
    """交互式术语提取器"""
    
    def __init__(self):
        self.extractor = None
        self.api_configured = self._check_api_config()
        # 调试输出
        if self.api_configured:
            print("✅ 检测到API配置")
        
    def _check_api_config(self) -> bool:
        """检查API配置是否有效"""
        try:
            from config import OPENAI_API_KEY
            # 更严格的验证：检查是否为空、是否为默认值、是否足够长
            is_valid = bool(
                OPENAI_API_KEY and 
                OPENAI_API_KEY.strip() != "" and
                OPENAI_API_KEY != "your-api-key-here" and 
                len(OPENAI_API_KEY.strip()) > 20  # API密钥通常很长
            )
            if is_valid:
                logger.debug(f"API密钥有效（长度: {len(OPENAI_API_KEY)}）")
            else:
                logger.debug(f"API密钥无效: 空={not OPENAI_API_KEY}, 长度={len(OPENAI_API_KEY) if OPENAI_API_KEY else 0}")
            return is_valid
        except ImportError as e:
            logger.debug(f"导入config失败: {e}")
            return False
        except Exception as e:
            logger.debug(f"API配置检查异常: {e}")
            return False
    
    def clear_screen(self):
        """清屏"""
        os.system('cls' if os.name == 'nt' else 'clear')
    
    def print_header(self):
        """打印标题"""
        print("\n" + "=" * 70)
        print("📚 期刊术语提取工具 - 交互式界面")
        print("=" * 70)
    
    def print_menu(self):
        """打印主菜单"""
        print("\n【主菜单】")
        print("  1. 📄 处理PDF文件（完整功能）")
        print("  2. 📁 批量处理文件夹中的所有PDF")
        print("  3. ⚙️  检查配置和依赖")
        print("  4. 📖 查看使用说明")
        print("  5. 🚪 退出程序")
        print()
    
    def check_dependencies(self):
        """检查依赖"""
        print("\n" + "=" * 70)
        print("🔍 检查系统依赖...")
        print("=" * 70)
        
        dependencies = {
            "requests": "requests (HTTP客户端)",
            "openpyxl": "openpyxl (Excel处理)",
            "openai": "OpenAI (GPT API客户端)",
            "tiktoken": "tiktoken (Token计数)",
            "pdfminer": "pdfminer.six (PDF文本提取)",
            "PyPDF2": "PyPDF2 (PDF处理)",
        }
        
        all_ok = True
        for module, desc in dependencies.items():
            try:
                __import__(module)
                print(f"  ✅ {desc}")
            except ImportError:
                print(f"  ❌ {desc} - 未安装")
                all_ok = False
        
        print()
        
        if not all_ok:
            print("⚠️  缺少部分依赖，请运行：")
            print("pip install -r requirements.txt")
        else:
            print("✅ 所有依赖已安装！")
        
        # 检查API配置
        print("\n" + "-" * 70)
        print("🔑 API配置检查...")
        print("-" * 70)
        
        # 检查OpenAI API
        try:
            from config import OPENAI_API_KEY, OPENAI_BASE_URL, DEFAULT_MODEL
            
            if OPENAI_API_KEY and OPENAI_API_KEY.strip() != "" and OPENAI_API_KEY != "your-api-key-here" and len(OPENAI_API_KEY.strip()) > 20:
                print(f"  ✅ OpenAI API密钥已配置（长度: {len(OPENAI_API_KEY)}）")
                print(f"  ✅ API端点: {OPENAI_BASE_URL}")
                print(f"  ✅ 使用模型: {DEFAULT_MODEL}")
                self.api_configured = True
            else:
                print(f"  ⚠️  OpenAI API密钥未配置或无效")
                if OPENAI_API_KEY:
                    print(f"  💡 当前API密钥长度: {len(OPENAI_API_KEY)}")
                print(f"  💡 请编辑 config.py 文件配置正确的API密钥")
                self.api_configured = False
        except ImportError:
            print(f"  ⚠️  config.py 文件不存在")
            self.api_configured = False
        except Exception as e:
            print(f"  ⚠️  检查API配置时出错: {e}")
            self.api_configured = False
        
        # 检查科大讯飞OCR配置
        print("\n" + "-" * 70)
        print("🔑 科大讯飞OCR配置检查...")
        print("-" * 70)
        
        try:
            from config import XUNFEI_OCR_CONFIG
            
            app_id = XUNFEI_OCR_CONFIG.get('app_id')
            secret = XUNFEI_OCR_CONFIG.get('secret')
            
            if app_id and secret and app_id != 'your-xunfei-app-id':
                print(f"  ✅ 科大讯飞OCR已配置（AppID: {app_id[:8]}***）")
                print(f"  ✅ 导出格式: {XUNFEI_OCR_CONFIG.get('export_format', 'txt')}")
            else:
                print(f"  ⚠️  科大讯飞OCR未配置")
                print(f"  💡 请编辑 config.py 中的 XUNFEI_OCR_CONFIG")
        except ImportError:
            print(f"  ⚠️  config.py 文件不存在")
        except Exception as e:
            print(f"  ⚠️  检查OCR配置时出错: {e}")
        
        print("=" * 70)
        input("\n按回车键返回主菜单...")
    
    def find_pdf_files(self, directory: str = ".") -> List[Path]:
        """查找PDF文件"""
        pdf_files = list(Path(directory).glob("*.pdf"))
        return sorted(pdf_files, key=lambda x: x.name)
    
    def select_pdf_file(self) -> Optional[Path]:
        """选择PDF文件"""
        pdf_files = self.find_pdf_files()
        
        if not pdf_files:
            print("\n❌ 当前目录未找到PDF文件")
            print("请将PDF文件放入以下目录：")
            print(f"  {Path('.').absolute()}")
            input("\n按回车键返回...")
            return None
        
        print("\n" + "=" * 70)
        print(f"📁 找到 {len(pdf_files)} 个PDF文件：")
        print("=" * 70)
        
        for idx, pdf in enumerate(pdf_files, 1):
            size_mb = pdf.stat().st_size / (1024 * 1024)
            print(f"  {idx}. {pdf.name} ({size_mb:.2f} MB)")
        
        print("\n  0. 返回主菜单")
        print()
        
        while True:
            try:
                choice = input("请选择文件编号: ").strip()
                
                if choice == "0":
                    return None
                
                idx = int(choice) - 1
                if 0 <= idx < len(pdf_files):
                    return pdf_files[idx]
                else:
                    print("❌ 无效的编号，请重试")
            except ValueError:
                print("❌ 请输入数字")
            except KeyboardInterrupt:
                print("\n")
                return None
    
    def process_single_pdf(self):
        """处理单个PDF文件"""
        pdf_file = self.select_pdf_file()
        
        if pdf_file is None:
            return
        
        print("\n" + "=" * 70)
        print(f"📄 准备处理: {pdf_file.name}")
        print("=" * 70)
        
        # 选择输出目录
        print("\n选择输出目录：")
        print("  1. PDF所在目录（默认）")
        print("  2. 指定目录")
        
        output_dir = None
        while True:
            choice = input("\n请选择 (1-2, 默认1): ").strip() or "1"
            
            if choice == "1":
                output_dir = pdf_file.parent
                break
            elif choice == "2":
                custom_dir = input("请输入输出目录路径: ").strip()
                if custom_dir:
                    output_dir = Path(custom_dir)
                    output_dir.mkdir(parents=True, exist_ok=True)
                    break
            else:
                print("❌ 无效选择")
        
        print(f"\n✅ 输出目录: {output_dir}")
        
        # 确认开始处理
        print("\n" + "-" * 70)
        if self.api_configured:
            print("✅ API已配置，将使用完整功能（关键词 + GPT摘要提取）")
        else:
            print("⚠️  注意: API未配置，将只提取关键词，不会使用GPT提取摘要术语")
            print("如需完整功能，请先配置 config.py 中的API密钥")
        
        confirm = input("\n是否开始处理? (Y/n): ").strip().lower()
        if confirm in ['n', 'no', '否']:
            print("已取消")
            input("\n按回车键返回...")
            return
        
        # 初始化提取器
        try:
            if self.extractor is None:
                print("\n🔧 正在初始化提取器...")
                self.extractor = JournalKeywordExtractor()
            
            # 处理PDF
            results = self.extractor.process_journal_pdf(
                pdf_path=str(pdf_file),
                output_dir=str(output_dir)
            )
            
            # 显示结果
            if results:
                print("\n" + "=" * 70)
                print("🎉 处理成功！生成的文件：")
                print("=" * 70)
                for key, file_path in results.items():
                    print(f"  📄 {Path(file_path).name}")
                print("=" * 70)
            else:
                print("\n⚠️  未生成任何文件，请查看上方错误信息")
            
        except KeyboardInterrupt:
            print("\n\n⏹️  用户中断处理")
        except Exception as e:
            print(f"\n❌ 处理失败: {e}")
            import traceback
            traceback.print_exc()
        
        input("\n按回车键返回主菜单...")
    
    def batch_process_pdfs(self):
        """批量处理PDF文件"""
        pdf_files = self.find_pdf_files()
        
        if not pdf_files:
            print("\n❌ 当前目录未找到PDF文件")
            input("\n按回车键返回...")
            return
        
        print("\n" + "=" * 70)
        print(f"📁 找到 {len(pdf_files)} 个PDF文件")
        print("=" * 70)
        
        for idx, pdf in enumerate(pdf_files, 1):
            size_mb = pdf.stat().st_size / (1024 * 1024)
            print(f"  {idx}. {pdf.name} ({size_mb:.2f} MB)")
        
        print("\n是否批量处理所有PDF文件?")
        if self.api_configured:
            print("✅ API已配置，将使用完整功能（关键词 + GPT摘要提取）")
        else:
            print("⚠️  注意: API未配置，将只提取关键词")
        
        confirm = input("\n确认开始批量处理? (Y/n): ").strip().lower()
        if confirm in ['n', 'no', '否']:
            print("已取消")
            input("\n按回车键返回...")
            return
        
        # 初始化提取器
        try:
            if self.extractor is None:
                print("\n🔧 正在初始化提取器...")
                self.extractor = JournalKeywordExtractor()
            
            # 批量处理
            success_count = 0
            failed_files = []
            
            for idx, pdf_file in enumerate(pdf_files, 1):
                print("\n" + "=" * 70)
                print(f"处理进度: {idx}/{len(pdf_files)}")
                print("=" * 70)
                
                try:
                    results = self.extractor.process_journal_pdf(
                        pdf_path=str(pdf_file)
                    )
                    
                    if results:
                        success_count += 1
                        print(f"✅ {pdf_file.name} - 处理成功")
                    else:
                        failed_files.append(pdf_file.name)
                        print(f"⚠️  {pdf_file.name} - 未生成文件")
                
                except KeyboardInterrupt:
                    print("\n\n⏹️  用户中断批量处理")
                    break
                except Exception as e:
                    failed_files.append(pdf_file.name)
                    print(f"❌ {pdf_file.name} - 处理失败: {e}")
            
            # 显示总结
            print("\n" + "=" * 70)
            print("📊 批量处理完成")
            print("=" * 70)
            print(f"  成功: {success_count}/{len(pdf_files)}")
            if failed_files:
                print(f"  失败: {len(failed_files)}")
                print("\n失败的文件:")
                for name in failed_files:
                    print(f"    - {name}")
            print("=" * 70)
            
        except Exception as e:
            print(f"\n❌ 批量处理异常: {e}")
            import traceback
            traceback.print_exc()
        
        input("\n按回车键返回主菜单...")
    
    def show_help(self):
        """显示使用说明"""
        print("\n" + "=" * 70)
        print("📖 使用说明")
        print("=" * 70)
        
        help_text = """
【功能介绍】
1. 智能PDF文本提取（文本型直接提取，扫描版使用科大讯飞OCR）
2. 自动识别"关键词"和"Key words"栏目
3. 匹配中英文关键词对，输出为Excel表格
4. 使用GPT-4o从摘要中智能提取术语对（需配置API）

【快速开始】
1. 将PDF文件放入程序所在目录
2. 运行此脚本（双击或 python main.py）
3. 选择"处理PDF文件"
4. 按提示操作

【命令行模式】
也可以通过命令行参数直接运行：
  python main.py <PDF文件> [选项]
  
选项：
  --output, -o     输出目录（默认为PDF所在目录）
  --api-key        OpenAI API密钥
  --base-url       API端点URL

【输出文件】
- XXX_关键词术语对.xlsx - 从关键词栏提取
- XXX_摘要术语对.xlsx - GPT从摘要提取（需API）

【API配置】
编辑 config.py 文件，配置以下内容：
  
  # OpenAI API（用于GPT术语提取）
  OPENAI_API_KEY = "your-openai-api-key"
  OPENAI_BASE_URL = "your-api-endpoint"
  
  # 科大讯飞OCR（用于扫描版PDF识别）
  XUNFEI_OCR_CONFIG = {
      "app_id": "your-xunfei-app-id",
      "secret": "your-xunfei-secret",
      ...
  }

【依赖安装】
如果缺少依赖，请运行：
  pip install -r requirements.txt

【注意事项】
1. 文本型PDF使用PyMuPDF直接提取（快速、免费）
2. 扫描版PDF使用科大讯飞OCR识别（需要网络和API）
3. API密钥请妥善保管，不要泄露
4. 建议在处理重要文件前先备份

【技术支持】
详细文档: README.md
日志文件: journal_extractor.log
"""
        print(help_text)
        print("=" * 70)
        input("\n按回车键返回主菜单...")
    
    def run(self):
        """运行交互式界面"""
        while True:
            try:
                self.clear_screen()
                self.print_header()
                self.print_menu()
                
                choice = input("请选择功能 (1-5): ").strip()
                
                if choice == "1":
                    self.process_single_pdf()
                elif choice == "2":
                    self.batch_process_pdfs()
                elif choice == "3":
                    self.check_dependencies()
                elif choice == "4":
                    self.show_help()
                elif choice == "5":
                    print("\n👋 感谢使用！再见！")
                    break
                else:
                    print("\n❌ 无效选择，请输入1-5")
                    input("按回车键继续...")
            
            except KeyboardInterrupt:
                print("\n\n👋 感谢使用！再见！")
                break
            except Exception as e:
                print(f"\n❌ 发生错误: {e}")
                import traceback
                traceback.print_exc()
                input("\n按回车键继续...")


# =============================================================================
# 主函数 - 支持两种运行模式
# =============================================================================

def main():
    """
    主函数 - 自动检测运行模式
    - 如果提供了命令行参数：使用命令行模式
    - 如果没有参数：启动交互式界面
    """
    import argparse
    
    # 创建参数解析器
    parser = argparse.ArgumentParser(
        description="期刊关键词和摘要术语提取工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:

【交互式模式】
  python main.py
  (不带任何参数时自动进入交互式菜单)

【命令行模式】
  python main.py 空天技术2024年2期_全.pdf
  python main.py 空天技术2024年2期_全.pdf --output results
  python main.py 空天技术2024年2期_全.pdf --api-key YOUR_KEY
        """
    )
    
    parser.add_argument("pdf_file", nargs='?', help="PDF文件路径（可选，不提供则进入交互式模式）")
    parser.add_argument("--output", "-o", help="输出目录（默认为PDF所在目录）")
    parser.add_argument("--api-key", help="OpenAI API密钥（用于GPT-4o摘要术语提取）")
    parser.add_argument("--base-url", help="OpenAI API基础URL")
    
    args = parser.parse_args()
    
    # 判断运行模式
    if args.pdf_file:
        # 命令行模式
        run_cli_mode(args)
    else:
        # 交互式模式
        run_interactive_mode()


def run_cli_mode(args):
    """命令行模式"""
    # 检查文件是否存在
    if not os.path.exists(args.pdf_file):
        print(f"❌ 文件不存在: {args.pdf_file}")
        sys.exit(1)
    
    # 获取API密钥
    api_key = args.api_key or os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("⚠️  未提供OpenAI API密钥")
        print("   - 将只提取关键词，不进行GPT摘要术语提取")
        print("   - 如需GPT功能，请使用 --api-key 参数或设置 OPENAI_API_KEY 环境变量")
        print()
    
    try:
        # 创建提取器
        extractor = JournalKeywordExtractor(
            api_key=api_key,
            base_url=args.base_url
        )
        
        # 处理PDF
        results = extractor.process_journal_pdf(
            pdf_path=args.pdf_file,
            output_dir=args.output
        )
        
        # 显示结果
        if results:
            print("✅ 处理成功！生成文件:")
            for key, file_path in results.items():
                print(f"   - {file_path}")
        else:
            print("⚠️  未生成任何输出文件")
            
    except KeyboardInterrupt:
        print("\n⏹️  用户中断程序")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 程序执行失败: {e}")
        logger.exception("程序异常")
        sys.exit(1)


def run_interactive_mode():
    """交互式模式"""
    try:
        app = InteractiveTermExtractor()
        app.run()
    except Exception as e:
        print(f"\n❌ 程序异常: {e}")
        import traceback
        traceback.print_exc()
        input("\n按回车键退出...")
        sys.exit(1)


if __name__ == "__main__":
    main()

