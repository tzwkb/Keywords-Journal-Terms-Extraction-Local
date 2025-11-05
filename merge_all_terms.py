#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
整合所有Excel文件中的术语对
"""

import os
import sys
from pathlib import Path
from typing import List, Dict, Set

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ 缺少openpyxl模块，请安装: pip install openpyxl")
    sys.exit(1)


class TermMerger:
    """术语整合器"""
    
    def __init__(self):
        self.all_keywords = []
        self.all_abstract_terms = []
        self.keyword_files = []
        self.abstract_files = []
    
    def find_excel_files(self):
        """查找所有术语Excel文件"""
        print("🔍 正在搜索所有术语Excel文件...")
        
        for file in os.listdir('.'):
            if file.endswith('.xlsx'):
                if '关键词术语对' in file:
                    self.keyword_files.append(file)
                elif '摘要术语对' in file:
                    self.abstract_files.append(file)
        
        self.keyword_files.sort()
        self.abstract_files.sort()
        
        print(f"✅ 找到 {len(self.keyword_files)} 个关键词术语对文件")
        print(f"✅ 找到 {len(self.abstract_files)} 个摘要术语对文件")
    
    def read_terms_from_excel(self, file_path: str) -> List[Dict[str, str]]:
        """从Excel文件中读取术语对"""
        terms = []
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 跳过标题行，从第2行开始读取
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 3:
                    zh_term = str(row[1]).strip() if row[1] else ""
                    en_term = str(row[2]).strip() if row[2] else ""
                    source = str(row[3]).strip() if len(row) > 3 and row[3] else Path(file_path).stem
                    
                    if zh_term or en_term:
                        terms.append({
                            'zh_term': zh_term,
                            'en_term': en_term,
                            'source': source,
                            'file': Path(file_path).stem
                        })
            
            wb.close()
            
        except Exception as e:
            print(f"   ⚠️  读取文件失败 {file_path}: {e}")
        
        return terms
    
    def merge_keywords(self):
        """整合所有关键词术语对"""
        print("\n" + "=" * 70)
        print("📑 整合关键词术语对...")
        print("=" * 70)
        
        for idx, file in enumerate(self.keyword_files, 1):
            print(f"   处理 [{idx}/{len(self.keyword_files)}]: {file}")
            terms = self.read_terms_from_excel(file)
            self.all_keywords.extend(terms)
            print(f"      提取了 {len(terms)} 个术语对")
        
        print(f"\n✅ 关键词术语对总计: {len(self.all_keywords)} 个")
    
    def merge_abstract_terms(self):
        """整合所有摘要术语对"""
        print("\n" + "=" * 70)
        print("📝 整合摘要术语对...")
        print("=" * 70)
        
        for idx, file in enumerate(self.abstract_files, 1):
            print(f"   处理 [{idx}/{len(self.abstract_files)}]: {file}")
            terms = self.read_terms_from_excel(file)
            self.all_abstract_terms.extend(terms)
            print(f"      提取了 {len(terms)} 个术语对")
        
        print(f"\n✅ 摘要术语对总计: {len(self.all_abstract_terms)} 个")
    
    def deduplicate_terms(self, terms: List[Dict[str, str]]) -> List[Dict[str, str]]:
        """去重术语对（基于中英文术语组合）"""
        seen = set()
        unique_terms = []
        
        for term in terms:
            # 创建唯一标识（忽略大小写）
            zh_lower = term['zh_term'].lower().strip()
            en_lower = term['en_term'].lower().strip()
            key = (zh_lower, en_lower)
            
            if key not in seen and (zh_lower or en_lower):
                seen.add(key)
                unique_terms.append(term)
        
        return unique_terms
    
    def save_merged_excel(self, terms: List[Dict[str, str]], output_file: str, 
                          title: str, header_color: str, term_type: str):
        """保存整合后的术语到Excel"""
        print(f"\n💾 正在保存到: {output_file}")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 写入标题行
        headers = ["序号", "中文术语", "英文术语", "来源类型", "来源文件"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        for idx, term in enumerate(terms, 1):
            row_data = [
                idx,
                term.get('zh_term', ''),
                term.get('en_term', ''),
                term.get('source', term_type),
                term.get('file', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=value)
                cell.border = border
                if col == 1:
                    cell.alignment = Alignment(horizontal="center")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 40
        
        # 保存文件
        wb.save(output_file)
        print(f"✅ 已保存 {len(terms)} 个术语对")
    
    def save_all_terms_combined(self, output_file: str):
        """保存所有术语（关键词+摘要）到一个文件"""
        print(f"\n💾 正在保存所有术语到: {output_file}")
        
        # 合并所有术语
        all_terms = []
        
        # 添加关键词术语
        for term in self.all_keywords:
            term_copy = term.copy()
            term_copy['category'] = '关键词'
            all_terms.append(term_copy)
        
        # 添加摘要术语
        for term in self.all_abstract_terms:
            term_copy = term.copy()
            term_copy['category'] = '摘要'
            all_terms.append(term_copy)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "所有术语对"
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="5B4DA0", end_color="5B4DA0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 写入标题行
        headers = ["序号", "中文术语", "英文术语", "术语类别", "来源类型", "来源文件"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        for idx, term in enumerate(all_terms, 1):
            row_data = [
                idx,
                term.get('zh_term', ''),
                term.get('en_term', ''),
                term.get('category', ''),
                term.get('source', ''),
                term.get('file', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=value)
                cell.border = border
                if col == 1:
                    cell.alignment = Alignment(horizontal="center")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 40
        
        # 保存文件
        wb.save(output_file)
        print(f"✅ 已保存 {len(all_terms)} 个术语对")
    
    def run(self):
        """运行整合流程"""
        print("\n" + "=" * 70)
        print("🎯 期刊术语整合工具")
        print("=" * 70)
        
        # 查找文件
        self.find_excel_files()
        
        if not self.keyword_files and not self.abstract_files:
            print("\n❌ 未找到任何术语Excel文件")
            return
        
        # 整合关键词术语
        if self.keyword_files:
            self.merge_keywords()
        
        # 整合摘要术语
        if self.abstract_files:
            self.merge_abstract_terms()
        
        # 保存结果
        print("\n" + "=" * 70)
        print("💾 保存整合结果...")
        print("=" * 70)
        
        # 1. 保存所有关键词术语（不去重）
        if self.all_keywords:
            self.save_merged_excel(
                self.all_keywords,
                "所有关键词术语对_完整版.xlsx",
                "关键词术语对",
                "366092",
                "文献关键词"
            )
        
        # 2. 保存所有摘要术语（不去重）
        if self.all_abstract_terms:
            self.save_merged_excel(
                self.all_abstract_terms,
                "所有摘要术语对_完整版.xlsx",
                "摘要术语对",
                "2E7D32",
                "GPT-4o摘要提取"
            )
        
        # 3. 保存去重后的关键词术语
        if self.all_keywords:
            unique_keywords = self.deduplicate_terms(self.all_keywords)
            print(f"\n📊 关键词术语去重: {len(self.all_keywords)} → {len(unique_keywords)}")
            self.save_merged_excel(
                unique_keywords,
                "所有关键词术语对_去重版.xlsx",
                "关键词术语对（去重）",
                "366092",
                "文献关键词"
            )
        
        # 4. 保存去重后的摘要术语
        if self.all_abstract_terms:
            unique_abstract = self.deduplicate_terms(self.all_abstract_terms)
            print(f"\n📊 摘要术语去重: {len(self.all_abstract_terms)} → {len(unique_abstract)}")
            self.save_merged_excel(
                unique_abstract,
                "所有摘要术语对_去重版.xlsx",
                "摘要术语对（去重）",
                "2E7D32",
                "GPT-4o摘要提取"
            )
        
        # 5. 保存所有术语合并（关键词+摘要）
        if self.all_keywords or self.all_abstract_terms:
            self.save_all_terms_combined("所有术语对_总表.xlsx")
        
        # 6. 保存所有术语去重版
        if self.all_keywords or self.all_abstract_terms:
            all_combined = []
            for term in self.all_keywords:
                term_copy = term.copy()
                term_copy['category'] = '关键词'
                all_combined.append(term_copy)
            for term in self.all_abstract_terms:
                term_copy = term.copy()
                term_copy['category'] = '摘要'
                all_combined.append(term_copy)
            
            unique_all = self.deduplicate_terms(all_combined)
            print(f"\n📊 所有术语去重: {len(all_combined)} → {len(unique_all)}")
            
            # 保存去重后的总表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "所有术语对（去重）"
            
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="5B4DA0", end_color="5B4DA0", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            headers = ["序号", "中文术语", "英文术语", "术语类别", "来源类型", "来源文件"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            for idx, term in enumerate(unique_all, 1):
                row_data = [
                    idx,
                    term.get('zh_term', ''),
                    term.get('en_term', ''),
                    term.get('category', ''),
                    term.get('source', ''),
                    term.get('file', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=idx + 1, column=col, value=value)
                    cell.border = border
                    if col == 1:
                        cell.alignment = Alignment(horizontal="center")
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 40
            
            wb.save("所有术语对_去重版.xlsx")
            print(f"💾 已保存到: 所有术语对_去重版.xlsx")
            print(f"✅ 已保存 {len(unique_all)} 个术语对")
        
        # 最终总结
        print("\n" + "=" * 70)
        print("🎉 整合完成！生成的文件：")
        print("=" * 70)
        print("📄 所有关键词术语对_完整版.xlsx - 关键词完整版（未去重）")
        print("📄 所有关键词术语对_去重版.xlsx - 关键词去重版")
        print("📄 所有摘要术语对_完整版.xlsx - 摘要术语完整版（未去重）")
        print("📄 所有摘要术语对_去重版.xlsx - 摘要术语去重版")
        print("📄 所有术语对_总表.xlsx - 关键词+摘要完整版（未去重）")
        print("📄 所有术语对_去重版.xlsx - 关键词+摘要去重版（推荐）")
        print("=" * 70)
        
        print(f"\n📊 统计信息：")
        print(f"   关键词术语: {len(self.all_keywords)} 个")
        print(f"   摘要术语: {len(self.all_abstract_terms)} 个")
        print(f"   总计: {len(self.all_keywords) + len(self.all_abstract_terms)} 个")
        print("=" * 70 + "\n")


def main():
    """主函数"""
    merger = TermMerger()
    merger.run()


if __name__ == "__main__":
    main()

